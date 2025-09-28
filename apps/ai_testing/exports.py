"""
AI Testing Data Export Functions
Comprehensive CSV/JSON export functionality for AI testing data
"""

import csv
import io
from datetime import timedelta
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment

from .models.test_coverage_gaps import TestCoverageGap, TestCoveragePattern
from .models.adaptive_thresholds import AdaptiveThreshold
from .models.regression_predictions import RegressionPrediction
from .dashboard_integration import get_ai_insights_summary


class AITestingExporter:
    """
    Main exporter class for AI testing data
    """

    def __init__(self, date_range_days=30):
        self.date_range_days = date_range_days
        self.since_date = timezone.now() - timedelta(days=date_range_days)

    def export_coverage_gaps_csv(self, filters=None, include_details=False):
        """
        Export coverage gaps as CSV

        Args:
            filters: Dictionary of filters to apply
            include_details: Include detailed information

        Returns:
            StringIO object containing CSV data
        """
        # Build queryset
        gaps = self._build_coverage_gaps_queryset(filters)

        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)

        # Headers
        headers = [
            'ID', 'Title', 'Coverage Type', 'Priority', 'Status',
            'Confidence Score', 'Impact Score', 'Effectiveness Score',
            'Affected Platforms', 'Affected Endpoints',
            'Recommended Framework', 'Estimated Time (hours)',
            'Similar Gaps Count', 'Identified At', 'Updated At'
        ]

        if include_details:
            headers.extend([
                'Description', 'Anomaly Type', 'Anomaly Severity',
                'Anomaly Occurrences', 'Assigned To',
                'Implementation File', 'Implementation Commit',
                'Verification Notes'
            ])

        writer.writerow(headers)

        # Data rows
        for gap in gaps:
            row = [
                str(gap.id),
                gap.title,
                gap.get_coverage_type_display(),
                gap.get_priority_display(),
                gap.get_status_display(),
                f"{gap.confidence_score:.3f}",
                f"{gap.impact_score:.1f}",
                f"{gap.effectiveness_score:.3f}",
                ', '.join(gap.affected_platforms) if gap.affected_platforms else '',
                ', '.join(gap.affected_endpoints) if gap.affected_endpoints else '',
                gap.get_recommended_framework_display() if gap.recommended_framework else '',
                gap.estimated_implementation_time,
                gap.similar_gaps_count,
                gap.identified_at.isoformat(),
                gap.updated_at.isoformat()
            ]

            if include_details:
                row.extend([
                    gap.description,
                    gap.anomaly_signature.anomaly_type if gap.anomaly_signature else '',
                    gap.anomaly_signature.severity if gap.anomaly_signature else '',
                    gap.anomaly_signature.occurrence_count if gap.anomaly_signature else 0,
                    gap.assigned_to.get_full_name() if gap.assigned_to else '',
                    gap.implemented_test_file or '',
                    gap.implementation_commit or '',
                    gap.verification_notes or ''
                ])

            writer.writerow(row)

        return output

    def export_coverage_gaps_json(self, filters=None, include_details=False):
        """
        Export coverage gaps as JSON

        Args:
            filters: Dictionary of filters to apply
            include_details: Include detailed information

        Returns:
            Dictionary containing structured JSON data
        """
        gaps = self._build_coverage_gaps_queryset(filters)

        export_data = {
            'export_info': {
                'generated_at': timezone.now().isoformat(),
                'total_count': gaps.count(),
                'date_range_days': self.date_range_days,
                'filters_applied': filters or {},
                'include_details': include_details
            },
            'coverage_gaps': []
        }

        for gap in gaps:
            gap_data = {
                'id': str(gap.id),
                'title': gap.title,
                'coverage_type': gap.coverage_type,
                'priority': gap.priority,
                'status': gap.status,
                'confidence_score': gap.confidence_score,
                'impact_score': gap.impact_score,
                'effectiveness_score': gap.effectiveness_score,
                'affected_platforms': gap.affected_platforms,
                'affected_endpoints': gap.affected_endpoints,
                'recommended_framework': gap.recommended_framework,
                'estimated_implementation_time': gap.estimated_implementation_time,
                'similar_gaps_count': gap.similar_gaps_count,
                'identified_at': gap.identified_at.isoformat(),
                'updated_at': gap.updated_at.isoformat()
            }

            if include_details:
                gap_data.update({
                    'description': gap.description,
                    'anomaly_signature': {
                        'id': str(gap.anomaly_signature.id) if gap.anomaly_signature else None,
                        'anomaly_type': gap.anomaly_signature.anomaly_type if gap.anomaly_signature else None,
                        'severity': gap.anomaly_signature.severity if gap.anomaly_signature else None,
                        'occurrence_count': gap.anomaly_signature.occurrence_count if gap.anomaly_signature else 0
                    },
                    'assigned_to': gap.assigned_to.get_full_name() if gap.assigned_to else None,
                    'pattern_metadata': gap.pattern_metadata,
                    'auto_generated_test_code': gap.auto_generated_test_code,
                    'implementation_details': {
                        'test_file': gap.implemented_test_file,
                        'commit': gap.implementation_commit,
                        'implemented_at': gap.implemented_at.isoformat() if gap.implemented_at else None,
                        'verified_at': gap.verified_at.isoformat() if gap.verified_at else None,
                        'verification_notes': gap.verification_notes
                    }
                })

            export_data['coverage_gaps'].append(gap_data)

        return export_data

    def export_coverage_gaps_excel(self, filters=None, include_details=False):
        """
        Export coverage gaps as Excel workbook

        Args:
            filters: Dictionary of filters to apply
            include_details: Include detailed information

        Returns:
            BytesIO object containing Excel workbook
        """
        gaps = self._build_coverage_gaps_queryset(filters)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Coverage Gaps"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            'ID', 'Title', 'Coverage Type', 'Priority', 'Status',
            'Confidence', 'Impact', 'Platforms', 'Framework',
            'Est. Time', 'Identified', 'Updated'
        ]

        if include_details:
            headers.extend(['Description', 'Anomaly Type', 'Assigned To'])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Write data
        for row, gap in enumerate(gaps, 2):
            data = [
                str(gap.id)[:8] + '...',  # Shortened ID
                gap.title,
                gap.get_coverage_type_display(),
                gap.get_priority_display(),
                gap.get_status_display(),
                f"{gap.confidence_score:.1%}",
                f"{gap.impact_score:.1f}/10",
                ', '.join(gap.affected_platforms[:2]) if gap.affected_platforms else '',
                gap.get_recommended_framework_display() if gap.recommended_framework else '',
                f"{gap.estimated_implementation_time}h",
                gap.identified_at.strftime('%Y-%m-%d'),
                gap.updated_at.strftime('%Y-%m-%d')
            ]

            if include_details:
                data.extend([
                    gap.description[:100] + '...' if len(gap.description) > 100 else gap.description,
                    gap.anomaly_signature.anomaly_type if gap.anomaly_signature else '',
                    gap.assigned_to.get_full_name() if gap.assigned_to else ''
                ])

            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def export_ai_insights_summary(self, format='json'):
        """
        Export comprehensive AI insights summary

        Args:
            format: Export format ('json', 'csv')

        Returns:
            Formatted export data
        """
        insights = get_ai_insights_summary()

        if format == 'json':
            return {
                'export_info': {
                    'generated_at': timezone.now().isoformat(),
                    'export_type': 'ai_insights_summary',
                    'date_range_days': self.date_range_days
                },
                'ai_insights': insights
            }

        elif format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Write summary data as key-value pairs
            writer.writerow(['Metric', 'Value', 'Category'])
            writer.writerow(['AI Health Score', insights['health_score'], 'Overall'])
            writer.writerow(['Total Coverage Gaps', insights['coverage_gaps']['total'], 'Coverage'])
            writer.writerow(['Critical Gaps', insights['coverage_gaps']['critical_count'], 'Coverage'])
            writer.writerow(['Regression Risk %', insights['regression_risk']['risk_score'], 'Risk'])
            writer.writerow(['Threshold Accuracy %', insights['threshold_status']['avg_accuracy'], 'Thresholds'])

            return output

    def export_patterns_analysis(self, format='json'):
        """
        Export pattern analysis data

        Args:
            format: Export format ('json', 'csv')

        Returns:
            Formatted export data
        """
        patterns = TestCoveragePattern.objects.filter(
            is_active=True,
            last_seen__gte=self.since_date
        ).order_by('-occurrence_count')

        if format == 'json':
            patterns_data = []
            for pattern in patterns:
                patterns_data.append({
                    'id': str(pattern.id),
                    'pattern_type': pattern.pattern_type,
                    'title': pattern.title,
                    'description': pattern.description,
                    'occurrence_count': pattern.occurrence_count,
                    'confidence_score': pattern.confidence_score,
                    'pattern_strength': pattern.pattern_strength,
                    'recommended_actions': pattern.recommended_actions,
                    'first_detected': pattern.first_detected.isoformat(),
                    'last_seen': pattern.last_seen.isoformat(),
                    'related_gaps_count': pattern.coverage_gaps.count()
                })

            return {
                'export_info': {
                    'generated_at': timezone.now().isoformat(),
                    'export_type': 'patterns_analysis',
                    'date_range_days': self.date_range_days
                },
                'patterns': patterns_data
            }

        elif format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow([
                'ID', 'Pattern Type', 'Title', 'Description',
                'Occurrence Count', 'Confidence Score', 'Pattern Strength',
                'Related Gaps', 'First Detected', 'Last Seen'
            ])

            # Data
            for pattern in patterns:
                writer.writerow([
                    str(pattern.id),
                    pattern.get_pattern_type_display(),
                    pattern.title,
                    pattern.description,
                    pattern.occurrence_count,
                    f"{pattern.confidence_score:.3f}",
                    f"{pattern.pattern_strength:.3f}",
                    pattern.coverage_gaps.count(),
                    pattern.first_detected.isoformat(),
                    pattern.last_seen.isoformat()
                ])

            return output

    def export_adaptive_thresholds(self, format='json'):
        """
        Export adaptive thresholds data

        Args:
            format: Export format ('json', 'csv')

        Returns:
            Formatted export data
        """
        thresholds = AdaptiveThreshold.objects.all().order_by('metric_name')

        if format == 'json':
            thresholds_data = []
            for threshold in thresholds:
                thresholds_data.append({
                    'id': str(threshold.id),
                    'metric_name': threshold.metric_name,
                    'value': threshold.value,
                    'confidence_interval': threshold.confidence_interval,
                    'sample_count': threshold.sample_count,
                    'accuracy': threshold.accuracy,
                    'precision': threshold.precision,
                    'created_at': threshold.created_at.isoformat(),
                    'updated_at': threshold.updated_at.isoformat()
                })

            return {
                'export_info': {
                    'generated_at': timezone.now().isoformat(),
                    'export_type': 'adaptive_thresholds',
                    'total_thresholds': len(thresholds_data)
                },
                'thresholds': thresholds_data
            }

        elif format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow([
                'Metric Name', 'Value', 'Confidence Interval',
                'Sample Count', 'Accuracy', 'Precision',
                'Updated At'
            ])

            # Data
            for threshold in thresholds:
                writer.writerow([
                    threshold.metric_name,
                    f"{threshold.value:.6f}",
                    f"{threshold.confidence_interval:.3f}",
                    threshold.sample_count,
                    f"{threshold.accuracy:.3f}",
                    f"{threshold.precision:.3f}",
                    threshold.updated_at.isoformat()
                ])

            return output

    def export_regression_predictions(self, format='json', limit=50):
        """
        Export regression predictions data

        Args:
            format: Export format ('json', 'csv')
            limit: Maximum number of predictions to export

        Returns:
            Formatted export data
        """
        predictions = RegressionPrediction.objects.filter(
            created_at__gte=self.since_date
        ).order_by('-created_at')[:limit]

        if format == 'json':
            predictions_data = []
            for prediction in predictions:
                predictions_data.append({
                    'id': str(prediction.id),
                    'version_identifier': prediction.version_identifier,
                    'risk_score': prediction.risk_score,
                    'risk_percentage': round(prediction.risk_score * 100, 1),
                    'confidence': prediction.confidence,
                    'risk_factors': prediction.risk_factors,
                    'prediction_metadata': prediction.prediction_metadata,
                    'created_at': prediction.created_at.isoformat()
                })

            return {
                'export_info': {
                    'generated_at': timezone.now().isoformat(),
                    'export_type': 'regression_predictions',
                    'date_range_days': self.date_range_days,
                    'predictions_count': len(predictions_data)
                },
                'predictions': predictions_data
            }

        elif format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # Headers
            writer.writerow([
                'ID', 'Version', 'Risk Score', 'Risk Percentage',
                'Confidence', 'Top Risk Factors', 'Created At'
            ])

            # Data
            for prediction in predictions:
                top_factors = ', '.join(
                    prediction.risk_factors.get('top_factors', [])[:3]
                ) if prediction.risk_factors else ''

                writer.writerow([
                    str(prediction.id),
                    prediction.version_identifier,
                    f"{prediction.risk_score:.4f}",
                    f"{prediction.risk_score * 100:.1f}%",
                    f"{prediction.confidence:.3f}",
                    top_factors,
                    prediction.created_at.isoformat()
                ])

            return output

    def export_comprehensive_report(self, format='json'):
        """
        Export comprehensive AI testing report

        Args:
            format: Export format ('json', 'excel')

        Returns:
            Comprehensive export data
        """
        if format == 'json':
            return {
                'export_info': {
                    'generated_at': timezone.now().isoformat(),
                    'export_type': 'comprehensive_ai_report',
                    'date_range_days': self.date_range_days
                },
                'ai_insights': get_ai_insights_summary(),
                'coverage_gaps': self.export_coverage_gaps_json(include_details=True),
                'patterns': self.export_patterns_analysis(format='json'),
                'thresholds': self.export_adaptive_thresholds(format='json'),
                'regression_predictions': self.export_regression_predictions(format='json')
            }

        elif format == 'excel':
            return self._create_comprehensive_excel_report()

    def _build_coverage_gaps_queryset(self, filters):
        """Build filtered queryset for coverage gaps"""
        gaps = TestCoverageGap.objects.select_related(
            'anomaly_signature', 'assigned_to'
        ).filter(identified_at__gte=self.since_date)

        if filters:
            if 'priority' in filters:
                gaps = gaps.filter(priority=filters['priority'])

            if 'status' in filters:
                gaps = gaps.filter(status=filters['status'])

            if 'coverage_type' in filters:
                gaps = gaps.filter(coverage_type=filters['coverage_type'])

            if 'min_confidence' in filters:
                gaps = gaps.filter(confidence_score__gte=filters['min_confidence'])

        return gaps.order_by('-confidence_score', '-impact_score')

    def _create_comprehensive_excel_report(self):
        """Create comprehensive Excel report with multiple sheets"""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Coverage Gaps Summary
        ws_gaps = wb.create_sheet("Coverage Gaps")
        self._create_gaps_excel_sheet(ws_gaps)

        # Sheet 2: Pattern Analysis
        ws_patterns = wb.create_sheet("Patterns")
        self._create_patterns_excel_sheet(ws_patterns)

        # Sheet 3: Thresholds
        ws_thresholds = wb.create_sheet("Thresholds")
        self._create_thresholds_excel_sheet(ws_thresholds)

        # Sheet 4: Summary Dashboard
        ws_summary = wb.create_sheet("AI Summary")
        self._create_summary_excel_sheet(ws_summary)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_gaps_excel_sheet(self, worksheet):
        """Create coverage gaps Excel sheet"""
        gaps = self._build_coverage_gaps_queryset(None)

        headers = [
            'Title', 'Type', 'Priority', 'Status', 'Confidence',
            'Impact', 'Platforms', 'Framework', 'Est. Hours', 'Identified'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data
        for row, gap in enumerate(gaps[:100], 2):  # Limit to 100 for Excel
            worksheet.cell(row=row, column=1, value=gap.title)
            worksheet.cell(row=row, column=2, value=gap.get_coverage_type_display())
            worksheet.cell(row=row, column=3, value=gap.get_priority_display())
            worksheet.cell(row=row, column=4, value=gap.get_status_display())
            worksheet.cell(row=row, column=5, value=f"{gap.confidence_score:.1%}")
            worksheet.cell(row=row, column=6, value=f"{gap.impact_score:.1f}")
            worksheet.cell(row=row, column=7, value=', '.join(gap.affected_platforms[:2]) if gap.affected_platforms else '')
            worksheet.cell(row=row, column=8, value=gap.get_recommended_framework_display() if gap.recommended_framework else '')
            worksheet.cell(row=row, column=9, value=gap.estimated_implementation_time)
            worksheet.cell(row=row, column=10, value=gap.identified_at.strftime('%Y-%m-%d'))

    def _create_patterns_excel_sheet(self, worksheet):
        """Create patterns Excel sheet"""
        patterns = TestCoveragePattern.objects.filter(
            is_active=True
        ).order_by('-occurrence_count')[:50]

        headers = [
            'Pattern Type', 'Title', 'Occurrences', 'Confidence',
            'Related Gaps', 'First Detected', 'Last Seen'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data
        for row, pattern in enumerate(patterns, 2):
            worksheet.cell(row=row, column=1, value=pattern.get_pattern_type_display())
            worksheet.cell(row=row, column=2, value=pattern.title)
            worksheet.cell(row=row, column=3, value=pattern.occurrence_count)
            worksheet.cell(row=row, column=4, value=f"{pattern.confidence_score:.1%}")
            worksheet.cell(row=row, column=5, value=pattern.coverage_gaps.count())
            worksheet.cell(row=row, column=6, value=pattern.first_detected.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=7, value=pattern.last_seen.strftime('%Y-%m-%d'))

    def _create_thresholds_excel_sheet(self, worksheet):
        """Create thresholds Excel sheet"""
        thresholds = AdaptiveThreshold.objects.all().order_by('metric_name')

        headers = [
            'Metric Name', 'Value', 'Accuracy', 'Precision',
            'Sample Count', 'Updated At'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Write data
        for row, threshold in enumerate(thresholds, 2):
            worksheet.cell(row=row, column=1, value=threshold.metric_name)
            worksheet.cell(row=row, column=2, value=f"{threshold.value:.6f}")
            worksheet.cell(row=row, column=3, value=f"{threshold.accuracy:.1%}")
            worksheet.cell(row=row, column=4, value=f"{threshold.precision:.1%}")
            worksheet.cell(row=row, column=5, value=threshold.sample_count)
            worksheet.cell(row=row, column=6, value=threshold.updated_at.strftime('%Y-%m-%d %H:%M'))

    def _create_summary_excel_sheet(self, worksheet):
        """Create AI summary Excel sheet"""
        insights = get_ai_insights_summary()

        # Summary data
        summary_data = [
            ['AI Testing Platform Summary', ''],
            ['Generated At', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Overall Health', ''],
            ['AI Health Score', f"{insights['health_score']}/100"],
            ['', ''],
            ['Coverage Gaps', ''],
            ['Total Gaps', insights['coverage_gaps']['total']],
            ['Critical Gaps', insights['coverage_gaps']['critical_count']],
            ['Recent (7d)', insights['coverage_gaps']['recent_7d']],
            ['', ''],
            ['Regression Risk', ''],
            ['Risk Score', f"{insights['regression_risk']['risk_score']}%"],
            ['Confidence', f"{insights['regression_risk']['confidence']}%"],
            ['', ''],
            ['Adaptive Thresholds', ''],
            ['Total Thresholds', insights['threshold_status']['total_thresholds']],
            ['Recent Updates', insights['threshold_status']['recent_updates']],
            ['Average Accuracy', f"{insights['threshold_status']['avg_accuracy']}%"],
        ]

        # Write data
        for row, (label, value) in enumerate(summary_data, 1):
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)

        # Style the summary
        for row in range(1, len(summary_data) + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value and not cell.value.strip():  # Empty rows
                continue
            elif cell.value in ['AI Testing Platform Summary', 'Overall Health', 'Coverage Gaps', 'Regression Risk', 'Adaptive Thresholds']:
                cell.font = Font(bold=True, size=12)


# Utility functions for export operations

def get_export_filename(data_type, format, timestamp=None):
    """Generate standardized export filename"""
    if not timestamp:
        timestamp = timezone.now()

    timestamp_str = timestamp.strftime('%Y%m%d_%H%M%S')
    return f"ai_testing_{data_type}_{timestamp_str}.{format}"


def validate_export_filters(filters):
    """Validate export filter parameters"""
    valid_filters = {}

    if 'priority' in filters:
        valid_priorities = ['critical', 'high', 'medium', 'low']
        if filters['priority'] in valid_priorities:
            valid_filters['priority'] = filters['priority']

    if 'status' in filters:
        valid_statuses = ['identified', 'test_generated', 'test_implemented', 'test_verified', 'dismissed']
        if filters['status'] in valid_statuses:
            valid_filters['status'] = filters['status']

    if 'coverage_type' in filters:
        valid_types = [
            'visual', 'performance', 'functional', 'integration',
            'edge_case', 'error_handling', 'user_flow', 'api_contract',
            'device_specific', 'network_condition'
        ]
        if filters['coverage_type'] in valid_types:
            valid_filters['coverage_type'] = filters['coverage_type']

    if 'min_confidence' in filters:
        try:
            confidence = float(filters['min_confidence'])
            if 0.0 <= confidence <= 1.0:
                valid_filters['min_confidence'] = confidence
        except (ValueError, TypeError):
            pass

    return valid_filters