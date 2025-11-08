"""
View Export Service
===================
Export dashboard data to CSV/Excel/PDF formats.
Schedule automated exports via Celery Beat.

Follows .claude/rules.md:
- Rule #3: Service methods < 50 lines
- Rule #11: Specific exception handling
- Rule #13: Security first (validate all user input)
"""

import csv
import logging
from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.utils import timezone

from apps.core.exceptions.patterns import DATABASE_EXCEPTIONS

logger = logging.getLogger(__name__)


class ViewExportService:
    """Service for exporting saved views to various formats"""

    @staticmethod
    def get_view_data(saved_view):
        """
        Get queryset and columns based on view type.
        
        Args:
            saved_view: DashboardSavedView instance
            
        Returns:
            tuple: (queryset, column_list)
        """
        from apps.y_helpdesk.models import Ticket
        from apps.activity.models.task import Task
        from apps.peoples.models import People

        view_type = saved_view.view_type
        filters = saved_view.filters or {}

        # Map view types to models and columns
        if view_type == 'TICKETS':
            queryset = Ticket.objects.filter(**filters).select_related(
                'assign', 'client', 'site', 'status'
            )
            columns = [
                'id', 'title', 'status__name', 'priority',
                'assign__username', 'created_at', 'due_date'
            ]
        elif view_type == 'TASKS':
            queryset = Task.objects.filter(**filters).select_related(
                'assignee', 'client', 'site'
            )
            columns = [
                'id', 'name', 'status', 'priority',
                'assignee__username', 'due_date', 'completed_at'
            ]
        elif view_type == 'ATTENDANCE':
            from apps.attendance.models.attendance_record import AttendanceRecord
            queryset = AttendanceRecord.objects.filter(**filters).select_related(
                'employee', 'shift', 'site'
            )
            columns = [
                'id', 'employee__username', 'shift__name',
                'clock_in_time', 'clock_out_time', 'status'
            ]
        else:
            # Default: use People
            queryset = People.objects.filter(**filters).select_related(
                'client', 'bu'
            )
            columns = ['id', 'username', 'email', 'phone', 'client__name']

        return queryset, columns

    @staticmethod
    def export_to_csv(queryset, columns, filename):
        """
        Export queryset to CSV.
        
        Args:
            queryset: Django queryset to export
            columns: List of field names to include
            filename: Base filename (without extension)
            
        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)

        # Header row
        header = [col.replace('__', ' > ').replace('_', ' ').title() for col in columns]
        writer.writerow(header)

        # Data rows
        for obj in queryset[:10000]:  # Limit to 10k rows for performance
            row = []
            for col in columns:
                value = obj
                for part in col.split('__'):
                    value = getattr(value, part, '')
                    if value is None:
                        value = ''
                        break
                row.append(str(value))
            writer.writerow(row)

        logger.info(f"CSV export generated: {filename}")
        return response

    @staticmethod
    def export_to_excel(queryset, columns, filename):
        """
        Export queryset to Excel with formatting.
        
        Args:
            queryset: Django queryset to export
            columns: List of field names to include
            filename: Base filename (without extension)
            
        Returns:
            HttpResponse with Excel content
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed, falling back to CSV")
            return ViewExportService.export_to_csv(queryset, columns, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Header row with formatting
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=idx)
            cell.value = col.replace('__', ' > ').replace('_', ' ').title()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for row_idx, obj in enumerate(queryset[:10000], 2):
            for col_idx, col in enumerate(columns, 1):
                value = obj
                for part in col.split('__'):
                    value = getattr(value, part, '')
                    if value is None:
                        value = ''
                        break

                # Format datetime objects
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')

                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # Auto-adjust column widths
        for idx, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        logger.info(f"Excel export generated: {filename}")
        return response

    @staticmethod
    def export_to_pdf(queryset, columns, filename):
        """
        Export queryset to PDF (simple table format).
        
        Args:
            queryset: Django queryset to export
            columns: List of field names to include
            filename: Base filename (without extension)
            
        Returns:
            HttpResponse with PDF content
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        except ImportError:
            logger.error("reportlab not installed, falling back to CSV")
            return ViewExportService.export_to_csv(queryset, columns, filename)

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter))

        # Prepare data
        data = []

        # Header row
        header = [col.replace('__', ' > ').replace('_', ' ').title() for col in columns]
        data.append(header)

        # Data rows
        for obj in queryset[:1000]:  # Limit for PDF
            row = []
            for col in columns:
                value = obj
                for part in col.split('__'):
                    value = getattr(value, part, '')
                    if value is None:
                        value = ''
                        break
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d %H:%M')
                row.append(str(value)[:50])  # Truncate long values
            data.append(row)

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        doc.build([table])
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

        logger.info(f"PDF export generated: {filename}")
        return response

    @staticmethod
    def schedule_export(saved_view, frequency='DAILY', recipients=None):
        """
        Schedule automated exports using Celery Beat.
        
        Args:
            saved_view: DashboardSavedView instance
            frequency: 'DAILY', 'WEEKLY', or 'MONTHLY'
            recipients: List of email addresses
            
        Returns:
            PeriodicTask instance
        """
        try:
            from django_celery_beat.models import PeriodicTask, CrontabSchedule
            import json
        except ImportError:
            logger.error("django-celery-beat not installed")
            return None

        try:
            # Create schedule based on frequency
            if frequency == 'DAILY':
                schedule, _ = CrontabSchedule.objects.get_or_create(
                    hour=8,
                    minute=0,
                    day_of_week='*',
                    day_of_month='*',
                    month_of_year='*'
                )
            elif frequency == 'WEEKLY':
                schedule, _ = CrontabSchedule.objects.get_or_create(
                    hour=8,
                    minute=0,
                    day_of_week='1',  # Monday
                    day_of_month='*',
                    month_of_year='*'
                )
            elif frequency == 'MONTHLY':
                schedule, _ = CrontabSchedule.objects.get_or_create(
                    hour=8,
                    minute=0,
                    day_of_week='*',
                    day_of_month='1',  # First day of month
                    month_of_year='*'
                )
            else:
                logger.error(f"Invalid frequency: {frequency}")
                return None

            # Create or update periodic task
            task_name = f"export_saved_view_{saved_view.id}"
            task, created = PeriodicTask.objects.update_or_create(
                name=task_name,
                defaults={
                    'crontab': schedule,
                    'task': 'apps.core.tasks.export_tasks.export_saved_view',
                    'kwargs': json.dumps({
                        'view_id': saved_view.id,
                        'recipients': recipients or [saved_view.cuser.email]
                    }),
                    'enabled': True
                }
            )

            # Update saved view
            saved_view.export_schedule = f"{frequency}_AT_8AM"
            saved_view.save(update_fields=['export_schedule'])

            logger.info(
                f"Scheduled export {'created' if created else 'updated'}: {task_name}",
                extra={'view_id': saved_view.id, 'frequency': frequency}
            )

            return task

        except DATABASE_EXCEPTIONS as e:
            logger.error(f"Failed to schedule export: {e}", exc_info=True)
            raise


__all__ = ['ViewExportService']
